"""
신차입금현황 엑셀 파서.

엑셀 구조 (시트 '상환스케줄'):
- 각 차입금은 6개 데이터 행 + 1 빈 행 구분자로 구성
- R(n)   A=대출명, B="회차",      C=빈,    D~=1,2,3,...
- R(n+1) A=빈,    B="납입일",   C=실행일, D~=회차별 납입일
- R(n+2) A=빈,    B="원금",     C=초기액, D~=회차별 원금
- R(n+3) A=빈,    B="이자",     C=빈,     D~=회차별 이자
- R(n+4) A=빈,    B="원리금",   C=빈,     D~=회차별 원리금
- R(n+5) A=빈,    B="미회수원금", C=초기, D~=회차별 미회수원금
- C 컬럼은 대출 실행일/초기액으로 스케줄 항목이 아님 → D 컬럼(idx 4)부터 수집
"""

from datetime import date, datetime
from io import BytesIO
from typing import Optional

import msoffcrypto
import openpyxl


SHEET_NAME = "상환스케줄"
DATA_START_COL = 4  # D 컬럼 (1-indexed)


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _to_decimal(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def parse_repayment_schedule(file_bytes: bytes, password: str = "7") -> list[dict]:
    """
    신차입금현황 엑셀 → 회차별 dict 리스트로 변환.

    반환 dict 키:
      block_index, loan_name, installment_no,
      original_due_date, principal, interest, total_payment, remaining_principal
    """
    # 비밀번호 복호화 시도 → 실패 시 일반 파일로 로드
    try:
        decrypted = BytesIO()
        office = msoffcrypto.OfficeFile(BytesIO(file_bytes))
        office.load_key(password=password)
        office.decrypt(decrypted)
        decrypted.seek(0)
        wb = openpyxl.load_workbook(decrypted, data_only=True)
    except Exception:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"시트 '{SHEET_NAME}'가 엑셀에 없습니다.")
    ws = wb[SHEET_NAME]

    rows: list[dict] = []
    block_index = 0

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if not a or b != "회차":
            continue

        # 블록 시작 발견
        block_index += 1
        loan_name = str(a).strip()

        # 헤더 행(r)에서 회차 번호들의 마지막 컬럼 위치 파악
        last_col = DATA_START_COL - 1
        for c in range(DATA_START_COL, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                last_col = c
        if last_col < DATA_START_COL:
            continue  # 회차 0건

        # 라벨 행 검증
        if ws.cell(r + 1, 2).value != "납입일":
            continue
        if ws.cell(r + 2, 2).value != "원금":
            continue
        if ws.cell(r + 3, 2).value != "이자":
            continue
        if ws.cell(r + 4, 2).value != "원리금":
            continue
        if ws.cell(r + 5, 2).value not in ("미회수 원금", "미회수원금"):
            continue

        for c in range(DATA_START_COL, last_col + 1):
            inst_no = ws.cell(r, c).value
            due_raw = ws.cell(r + 1, c).value
            due = _to_date(due_raw)
            if inst_no is None or due is None:
                continue
            try:
                inst_no_int = int(inst_no)
            except (ValueError, TypeError):
                continue

            rows.append({
                "block_index": block_index,
                "loan_name": loan_name,
                "installment_no": inst_no_int,
                "original_due_date": due,
                "principal": _to_decimal(ws.cell(r + 2, c).value),
                "interest": _to_decimal(ws.cell(r + 3, c).value),
                "total_payment": _to_decimal(ws.cell(r + 4, c).value),
                "remaining_principal": _to_decimal(ws.cell(r + 5, c).value),
            })

    return rows


def _parse_maturity_from_bigo(bigo: str) -> Optional[date]:
    """비고 컬럼에서 만기일 추출: "시작~종료(설명)" 또는 "YYYY.MM.DD" 형식"""
    if not bigo:
        return None
    s = str(bigo).strip()
    if '~' in s:
        after = s.split('~', 1)[1].strip()
        for sep in ['(', ',', ' ']:
            if sep in after:
                after = after.split(sep)[0].strip()
        s = after
    else:
        s = s.split('(')[0].split(',')[0].strip()
    parts = s.split('.')
    try:
        if len(parts) == 3:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        pass
    return None


def parse_financial_products(file_bytes: bytes) -> list[dict]:
    """
    금융상품 엑셀 파싱.

    지원 형식:
      헤더행: 은행 | 계좌명 | 상품명 | 전일잔액 | 증가 | 감소 | 당일잔액 | 비고
      (헤더 순서는 컬럼명으로 자동 탐색)

    반환 dict 키:
      product_code, company_name, product_name, amount, original_maturity_date
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {e}")

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 헤더 행 탐색 (은행/비고 컬럼이 있는 행)
        header_row = None
        col_map: dict[str, int] = {}
        for r in range(1, min(ws.max_row + 1, 30)):
            row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if '은행' in row_vals and '비고' in row_vals:
                header_row = r
                for idx, val in enumerate(row_vals, start=1):
                    col_map[val] = idx
                break

        if not header_row or '은행' not in col_map or '비고' not in col_map:
            continue

        c_company = col_map.get('은행')
        c_account = col_map.get('계좌명')
        c_product = col_map.get('상품명')
        c_amount  = col_map.get('당일잔액') or col_map.get('전일잔액')
        c_bigo    = col_map.get('비고')

        for r in range(header_row + 1, ws.max_row + 1):
            company = str(ws.cell(r, c_company).value or '').strip()
            if not company or company in ('소계', '합계', ''):
                continue

            account  = str(ws.cell(r, c_account).value or '').strip() if c_account else ''
            prod_raw = str(ws.cell(r, c_product).value or '').strip() if c_product else ''
            product  = prod_raw if prod_raw and prod_raw != '-' else account

            amount_raw = ws.cell(r, c_amount).value if c_amount else None
            amount = _to_decimal(amount_raw)
            if amount is None:
                continue

            bigo_val = ws.cell(r, c_bigo).value
            # 날짜 객체로 직접 들어올 수도 있음
            if isinstance(bigo_val, (date, datetime)):
                maturity = bigo_val.date() if isinstance(bigo_val, datetime) else bigo_val
            else:
                maturity = _parse_maturity_from_bigo(str(bigo_val or ''))

            if not maturity:
                continue

            product_code = f"{company}_{account}_{prod_raw}".replace(' ', '')
            results.append({
                "product_code": product_code,
                "company_name": company,
                "product_name": product,
                "amount": amount,
                "original_maturity_date": maturity,
            })

    return results
